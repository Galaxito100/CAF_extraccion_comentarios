import streamlit as st
import tempfile
import os
import zipfile
import io
from docx import Document
from lxml import etree
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import fitz
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
from reportlab.lib.units import cm
from reportlab.lib import colors


# ─────────────────────────────────────────────
# CONFIGURACIÓN DE PÁGINA
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Extractor de Comentarios | CAF",
    page_icon="💬",
    layout="centered"
)

# CSS corporativo CAF
st.markdown("""
<style>
    /* Fondo general */
    .stApp {
        background-color: #F5F7FA;
    }

    /* Ocultar header de Streamlit */
    header[data-testid="stHeader"] {
        background-color: #1B3A6B;
    }

    /* Barra superior CAF */
    .caf-topbar {
        background-color: #1B3A6B;
        padding: 0px 0px 0px 0px;
        margin: -60px -60px 0px -60px;
        height: 8px;
    }
    .caf-greenbar {
        background-color: #4CAF50;
        height: 5px;
        margin: 0px -60px 30px -60px;
    }

    /* Header principal */
    .caf-header {
        background: linear-gradient(135deg, #1B3A6B 0%, #2C5F8A 100%);
        border-radius: 12px;
        padding: 30px 40px;
        margin-bottom: 30px;
        text-align: center;
        box-shadow: 0 4px 15px rgba(27,58,107,0.3);
    }
    .caf-header h1 {
        color: white;
        font-size: 2rem;
        font-weight: 700;
        margin: 0 0 8px 0;
        letter-spacing: 1px;
    }
    .caf-header .subtitulo {
        color: #A8C8E8;
        font-size: 0.95rem;
        margin: 0;
    }
    .caf-header .division {
        color: #4CAF50;
        font-size: 0.85rem;
        font-weight: 600;
        margin-top: 8px;
        letter-spacing: 2px;
        text-transform: uppercase;
    }

    /* Franja verde debajo del header */
    .caf-accent {
        height: 4px;
        background: linear-gradient(90deg, #4CAF50, #81C784);
        border-radius: 2px;
        margin-bottom: 25px;
    }

    /* Tarjetas de sección */
    .caf-card {
        background: white;
        border-radius: 10px;
        padding: 25px 30px;
        margin-bottom: 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        border-left: 4px solid #4CAF50;
    }
    .caf-card h3 {
        color: #1B3A6B;
        margin: 0 0 15px 0;
        font-size: 1rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 1px;
    }

    /* Botón principal */
    .stButton > button {
        background: linear-gradient(135deg, #1B3A6B, #2C5F8A) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 12px 35px !important;
        font-size: 1rem !important;
        font-weight: 600 !important;
        width: 100% !important;
        transition: all 0.3s !important;
        box-shadow: 0 3px 10px rgba(27,58,107,0.3) !important;
    }
    .stButton > button:hover {
        background: linear-gradient(135deg, #4CAF50, #2E7D32) !important;
        box-shadow: 0 5px 15px rgba(76,175,80,0.4) !important;
        transform: translateY(-1px) !important;
    }

    /* Botones de descarga */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #4CAF50, #2E7D32) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 10px 25px !important;
        font-weight: 600 !important;
        width: 100% !important;
        box-shadow: 0 3px 10px rgba(76,175,80,0.3) !important;
    }

    /* Multiselect */
    .stMultiSelect > div > div {
        border-color: #4CAF50 !important;
        border-radius: 8px !important;
    }

    /* File uploader */
    [data-testid="stFileUploader"] {
        border: 2px dashed #4CAF50 !important;
        border-radius: 10px !important;
        background: #F0F7F0 !important;
    }

    /* Success / warning / info */
    .stSuccess {
        background-color: #E8F5E9 !important;
        border-left: 4px solid #4CAF50 !important;
        border-radius: 6px !important;
    }

    /* Footer */
    .caf-footer {
        text-align: center;
        color: #9E9E9E;
        font-size: 0.8rem;
        margin-top: 40px;
        padding-top: 20px;
        border-top: 1px solid #E0E0E0;
    }
    .caf-footer span {
        color: #4CAF50;
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# FUNCIONES (igual que antes)
# ─────────────────────────────────────────────

def extraer_celda_siguiente(tabla, label):
    for fila in tabla.rows:
        for j, celda in enumerate(fila.cells):
            if label.lower() in celda.text.strip().lower():
                if j + 1 < len(fila.cells):
                    valor = fila.cells[j + 1].text.strip()
                    if valor and valor.lower() != label.lower():
                        return valor
    return "NA"

def extraer_casilla_marcada(tabla, label):
    NS = {
        'w':   'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
        'w14': 'http://schemas.microsoft.com/office/word/2010/wordml',
    }
    MARCAS_UNICODE = {'☒', '☑'}
    for fila in tabla.rows:
        texto_fila = ' '.join(c.text for c in fila.cells).lower()
        if label.lower() not in texto_fila:
            continue
        resultados = []
        for celda in fila.cells:
            xml_celda = celda._tc
            for parrafo in xml_celda.findall('.//w:p', NS):
                runs = parrafo.findall('.//w:r', NS)
                for idx, run in enumerate(runs):
                    checked_elem = run.find('.//w14:checked', NS)
                    if checked_elem is not None:
                        val = checked_elem.get('{http://schemas.microsoft.com/office/word/2010/wordml}val', '0')
                        if val == '1':
                            for run_sig in runs[idx + 1:]:
                                t = run_sig.find('w:t', NS)
                                if t is not None and t.text and t.text.strip():
                                    if t.text.strip() not in MARCAS_UNICODE:
                                        resultados.append(t.text.strip())
                                    break
                    t_elem = run.find('w:t', NS)
                    if t_elem is not None and t_elem.text:
                        for char in t_elem.text:
                            if char in MARCAS_UNICODE:
                                resto = t_elem.text[t_elem.text.index(char) + 1:].strip()
                                if resto:
                                    resultados.append(resto)
                                elif idx + 1 < len(runs):
                                    t_sig = runs[idx + 1].find('w:t', NS)
                                    if t_sig is not None and t_sig.text:
                                        texto_sig = t_sig.text.strip()
                                        if texto_sig and texto_sig not in MARCAS_UNICODE:
                                            resultados.append(texto_sig)
                    fld = run.find('.//w:fldChar', NS)
                    if fld is not None:
                        fld_type = fld.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}fldCharType', '')
                        checked  = fld.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}checked', '0')
                        if fld_type == 'begin' and checked == '1':
                            for run_sig in runs[idx + 1:]:
                                t = run_sig.find('w:t', NS)
                                if t is not None and t.text and t.text.strip():
                                    resultados.append(t.text.strip())
                                    break
        if resultados:
            return ', '.join(dict.fromkeys(resultados))
    return "NA"

def extraer_metadatos(docx_path):
    doc = Document(docx_path)
    metadatos = {
        'operacion': 'NA', 'pais': 'NA', 'cliente': 'NA', 'garante': 'NA',
        'organismo_ejecutor': 'NA', 'sector_institucional': 'NA',
        'riesgo_institucional': 'Soberano', 'instancia_aprobatoria': 'NA',
        'unidad_negocio': 'NA', 'ejecutivo_pais': 'NA',
    }
    for tabla in doc.tables:
        texto_tabla = ' '.join(celda.text for fila in tabla.rows for celda in fila.cells).lower()
        palabras_clave = ['operación', 'operacion', 'país', 'pais', 'cliente',
                          'garante', 'ejecutivo', 'instancia', 'unidad de negocio']
        if not any(p in texto_tabla for p in palabras_clave):
            continue
        for campo, labels in [
            ('operacion',          ['Operación', 'Operacion']),
            ('pais',               ['País', 'Pais']),
            ('cliente',            ['Cliente']),
            ('garante',            ['Garante']),
            ('organismo_ejecutor', ['Organismo ejecutor', 'Organismo Ejecutor']),
            ('unidad_negocio',     ['Unidad de negocio', 'Unidad de Negocio']),
            ('ejecutivo_pais',     ['Ejecutivo/a país', 'Ejecutivo/a pais', 'Ejecutivo/a País']),
        ]:
            if metadatos[campo] == 'NA':
                for label in labels:
                    valor = extraer_celda_siguiente(tabla, label)
                    if valor and valor != 'NA':
                        metadatos[campo] = valor
                        break
        for campo, label in [
            ('sector_institucional',  'Sector institucional'),
            ('instancia_aprobatoria', 'Instancia aprobatoria'),
        ]:
            if metadatos[campo] == 'NA':
                valor = extraer_casilla_marcada(tabla, label)
                if valor != 'NA':
                    metadatos[campo] = valor
    return metadatos

def extraer_comentarios_word(docx_path):
    comentarios = []
    with zipfile.ZipFile(docx_path, 'r') as z:
        if 'word/comments.xml' not in z.namelist():
            return comentarios
        ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
        xml_comments  = z.read('word/comments.xml')
        tree_comments = etree.fromstring(xml_comments)
        datos_comentarios = {}
        ids_respuesta     = set()
        for comment in tree_comments.findall('.//w:comment', ns):
            cid         = comment.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}id')
            autor       = comment.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}author', 'Desconocido')
            fecha       = comment.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}date', '')
            fecha_corta = fecha[:10] if fecha else 'Sin fecha'
            textos      = comment.findall('.//w:t', ns)
            texto       = ' '.join(t.text for t in textos if t.text)
            if texto.strip():
                datos_comentarios[cid] = {'autor': autor, 'fecha': fecha_corta, 'texto': texto.strip()}
        comentarios_con_respuesta = {}
        if 'word/commentsExtended.xml' in z.namelist():
            xml_ext  = z.read('word/commentsExtended.xml')
            tree_ext = etree.fromstring(xml_ext)
            ns_ext   = {'w15': 'http://schemas.microsoft.com/office/word/2012/wordml'}
            xml_doc  = z.read('word/document.xml')
            tree_doc = etree.fromstring(xml_doc)
            para_id_a_cid = {}
            for elem in tree_doc.iter():
                tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                if tag == 'commentRangeStart':
                    cid_val = elem.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}id')
                    para_id = elem.get('{http://schemas.microsoft.com/office/word/2010/wordml}paraId')
                    if para_id and cid_val:
                        para_id_a_cid[para_id] = cid_val
            for ext in tree_ext.findall('.//w15:commentEx', ns_ext):
                para_id_hijo  = ext.get('{http://schemas.microsoft.com/office/word/2012/wordml}paraId')
                para_id_padre = ext.get('{http://schemas.microsoft.com/office/word/2012/wordml}paraIdParent')
                done          = ext.get('{http://schemas.microsoft.com/office/word/2012/wordml}done', '0')
                if para_id_padre:
                    ids_respuesta.add(para_id_hijo)
                    if para_id_padre in para_id_a_cid:
                        cid_padre_real  = para_id_a_cid[para_id_padre]
                        autor_respuesta = next(
                            (d['autor'] for c, d in datos_comentarios.items() if c == para_id_hijo), 'Desconocido'
                        )
                        comentarios_con_respuesta[cid_padre_real] = {
                            'respondido_por': autor_respuesta, 'resuelto': done == '1'
                        }
        xml_doc  = z.read('word/document.xml')
        tree_doc = etree.fromstring(xml_doc)
        body     = tree_doc.find('.//w:body', ns)
        parrafo_por_comentario = {}
        for parrafo in body.findall('.//w:p', ns):
            textos_p = parrafo.findall('.//w:t', ns)
            texto_p  = ''.join(t.text for t in textos_p if t.text).strip()
            if texto_p:
                for elem in parrafo.iter():
                    tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                    if tag == 'commentRangeStart':
                        cid_val = elem.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}id')
                        if cid_val:
                            parrafo_por_comentario[cid_val] = texto_p
        textos_referenciados = {cid: [] for cid in datos_comentarios}
        capturando           = {cid: False for cid in datos_comentarios}
        for elem in body.iter():
            tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            cid = elem.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}id')
            if tag == 'commentRangeStart' and cid in capturando:
                capturando[cid] = True
            if tag == 't' and elem.text:
                for c in capturando:
                    if capturando[c]:
                        textos_referenciados[c].append(elem.text)
            if tag == 'commentRangeEnd' and cid in capturando:
                capturando[cid] = False
    for cid, datos in datos_comentarios.items():
        if cid in ids_respuesta:
            continue
        ref_fragmento    = ''.join(textos_referenciados.get(cid, []))
        parrafo_completo = parrafo_por_comentario.get(cid, '(párrafo no encontrado)')
        if cid in comentarios_con_respuesta:
            info           = comentarios_con_respuesta[cid]
            estado         = 'Resuelto' if info['resuelto'] else 'En revisión'
            respondido_por = info['respondido_por']
        else:
            estado         = 'Pendiente'
            respondido_por = ''
        comentarios.append({
            'autor': datos['autor'], 'fecha': datos['fecha'], 'texto': datos['texto'],
            'referencia':       ref_fragmento.strip() if ref_fragmento.strip() else '(sin fragmento)',
            'parrafo_completo': parrafo_completo,
            'estado':           estado,
            'respondido_por':   respondido_por,
            'tipo_archivo':     'Word',
        })
    return comentarios

def extraer_comentarios_pdf(pdf_path):
    comentarios = []
    doc = fitz.open(pdf_path)
    for num_pagina, pagina in enumerate(doc, start=1):
        for annot in pagina.annots():
            texto     = annot.info.get('content', '').strip()
            autor     = annot.info.get('title', 'Desconocido')
            fecha_raw = annot.info.get('modDate', '') or annot.info.get('creationDate', '')
            fecha_corta = 'Sin fecha'
            if fecha_raw and len(fecha_raw) >= 10:
                try:
                    fecha_corta = f"{fecha_raw[2:6]}-{fecha_raw[6:8]}-{fecha_raw[8:10]}"
                except Exception:
                    pass
            if not texto:
                continue
            try:
                referencia = pagina.get_text("text", clip=annot.rect).strip()
            except Exception:
                referencia = '(sin fragmento)'
            comentarios.append({
                'autor': autor, 'fecha': fecha_corta, 'texto': texto,
                'referencia':       referencia if referencia else '(sin fragmento)',
                'parrafo_completo': f'Página {num_pagina}',
                'estado':           'Pendiente',
                'respondido_por':   '',
                'tipo_archivo':     'PDF',
            })
    doc.close()
    return comentarios

def generar_excel_bytes(comentarios):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comentarios"
    encabezado_font  = Font(bold=True, color="FFFFFF")
    encabezado_fill  = PatternFill("solid", fgColor="1B3A6B")
    encabezado_align = Alignment(horizontal="center", vertical="center")
    columnas = [
        ("Documento", 25), ("Tipo", 10), ("Comentario", 60), ("Fecha", 15),
        ("Autor", 25), ("Frase referenciada", 60), ("Texto Referenciado", 60),
        ("Respondido por", 25), ("Operación", 40), ("País", 15), ("Cliente", 25),
        ("Garante", 15), ("Organismo Ejecutor", 25), ("Sector Institucional", 20),
        ("Riesgo Institucional", 20), ("Instancia Aprobatoria", 30),
        ("Unidad de Negocio", 35), ("Ejecutivo/a País", 25),
    ]
    for col, (titulo, ancho) in enumerate(columnas, start=1):
        letra = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[letra].width = ancho
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font      = encabezado_font
        celda.fill      = encabezado_fill
        celda.alignment = encabezado_align
    for i, c in enumerate(comentarios, start=2):
        m = c.get('metadatos', {})
        valores = [
            c.get('documento', ''), c.get('tipo_archivo', ''), c['texto'],
            c['fecha'], c['autor'], c['referencia'], c['parrafo_completo'],
            c['respondido_por'],
            m.get('operacion', ''), m.get('pais', ''), m.get('cliente', ''),
            m.get('garante', ''), m.get('organismo_ejecutor', ''),
            m.get('sector_institucional', ''), m.get('riesgo_institucional', ''),
            m.get('instancia_aprobatoria', ''), m.get('unidad_negocio', ''),
            m.get('ejecutivo_pais', ''),
        ]
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=i, column=col, value=valor)
            celda.alignment = Alignment(vertical="top", wrap_text=True)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generar_pdf_bytes(comentarios):
    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm,   bottomMargin=2*cm
    )
    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'Titulo', parent=styles['Title'],
        fontSize=16, textColor=colors.HexColor('#1B3A6B'), spaceAfter=6
    )
    estilo_subtitulo = ParagraphStyle(
        'Subtitulo', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#4CAF50'),
        spaceAfter=20, fontName='Helvetica-Bold'
    )
    estilo_doc = ParagraphStyle(
        'Documento', parent=styles['Heading1'],
        fontSize=12, textColor=colors.white,
        backColor=colors.HexColor('#1B3A6B'),
        spaceAfter=10, spaceBefore=20,
        leftIndent=0, borderPadding=(6, 10, 6, 10),
    )
    estilo_meta = ParagraphStyle(
        'Meta', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#7F8C8D'), spaceAfter=4
    )
    estilo_ref = ParagraphStyle(
        'Ref', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#555555'),
        backColor=colors.HexColor('#F0F7F0'),
        leftIndent=10, rightIndent=10, spaceAfter=6, leading=14
    )
    estilo_texto = ParagraphStyle(
        'Texto', parent=styles['Normal'],
        fontSize=11, textColor=colors.HexColor('#1B3A6B'),
        spaceAfter=6, leading=15
    )
    contenido = []
    contenido.append(Paragraph("Extractor de Comentarios", estilo_titulo))
    contenido.append(Paragraph("DIRECCIÓN DE RIESGO SOBERANO — CAF", estilo_subtitulo))
    contenido.append(HRFlowable(width="100%", thickness=3, color=colors.HexColor('#4CAF50')))
    contenido.append(Spacer(1, 0.3*cm))
    contenido.append(Paragraph(f"Total de comentarios: {len(comentarios)}", styles['Normal']))
    contenido.append(Spacer(1, 0.4*cm))

    documento_actual = None
    contador = 1
    for c in comentarios:
        doc_nombre = c.get('documento', 'Sin nombre')
        if doc_nombre != documento_actual:
            documento_actual = doc_nombre
            contenido.append(Spacer(1, 0.3*cm))
            contenido.append(Paragraph(f"  {doc_nombre}", estilo_doc))
            contenido.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#4CAF50')))
            contenido.append(Spacer(1, 0.2*cm))
        contenido.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#E0E0E0')))
        contenido.append(Spacer(1, 0.2*cm))
        contenido.append(Paragraph(
            f"<b>Comentario #{contador}</b> — {c['autor']} · {c['fecha']} · {c['estado']}",
            estilo_meta
        ))
        contenido.append(Paragraph(
            f"<i>Texto referenciado:</i> \"{c['referencia']}\"", estilo_ref
        ))
        contenido.append(Paragraph(c['texto'], estilo_texto))
        if c['respondido_por']:
            contenido.append(Paragraph(
                f"<i>Respondido por:</i> {c['respondido_por']}", estilo_meta
            ))
        contenido.append(Spacer(1, 0.3*cm))
        contador += 1
    doc.build(contenido)
    buffer.seek(0)
    return buffer.getvalue()


# ─────────────────────────────────────────────
# INTERFAZ
# ─────────────────────────────────────────────

# Header
st.markdown("""
<div class="caf-topbar"></div>
<div class="caf-greenbar"></div>
<div class="caf-header">
    <h1>💬 Extractor de Comentarios</h1>
    <p class="subtitulo">Procesá documentos Word y PDF para extraer y consolidar comentarios</p>
    <p class="division">Dirección de Riesgo Soberano · CAF</p>
</div>
<div class="caf-accent"></div>
""", unsafe_allow_html=True)

# Sección subida
st.markdown('<div class="caf-card"><h3>📁 Archivos</h3>', unsafe_allow_html=True)
archivos_subidos = st.file_uploader(
    "Seleccioná los archivos",
    type=["docx", "pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed"
)
st.markdown('</div>', unsafe_allow_html=True)

# Sección formato
st.markdown('<div class="caf-card"><h3>📄 Formato de salida</h3>', unsafe_allow_html=True)
output_elegido = st.multiselect(
    "¿Qué deseas generar?",
    options=["Excel", "PDF"],
    default=["Excel"],
    label_visibility="visible"
)
st.markdown('</div>', unsafe_allow_html=True)

# Botón procesar
procesar = st.button("⚙️  Procesar documentos", disabled=not archivos_subidos)

if procesar:
    todos_los_comentarios = []
    progress = st.progress(0, text="Iniciando...")

    for idx, archivo in enumerate(archivos_subidos):
        nombre    = os.path.splitext(archivo.name)[0]
        extension = os.path.splitext(archivo.name)[1].lower()
        progress.progress((idx) / len(archivos_subidos), text=f"Procesando {archivo.name}...")

        with tempfile.NamedTemporaryFile(delete=False, suffix=extension) as tmp:
            tmp.write(archivo.read())
            tmp_path = tmp.name

        try:
            if extension == '.docx':
                metadatos   = extraer_metadatos(tmp_path)
                comentarios = extraer_comentarios_word(tmp_path)
            else:
                metadatos = {k: 'NA' for k in [
                    'operacion', 'pais', 'cliente', 'garante', 'organismo_ejecutor',
                    'sector_institucional', 'riesgo_institucional',
                    'instancia_aprobatoria', 'unidad_negocio', 'ejecutivo_pais'
                ]}
                metadatos['riesgo_institucional'] = 'Soberano'
                comentarios = extraer_comentarios_pdf(tmp_path)

            for c in comentarios:
                c['documento'] = nombre
                c['metadatos'] = metadatos
            todos_los_comentarios.extend(comentarios)
            st.success(f"✅ **{archivo.name}** — {len(comentarios)} comentarios encontrados")

        except Exception as e:
            st.warning(f"⚠️ Error en **{archivo.name}**: {e}")
        finally:
            os.unlink(tmp_path)

        progress.progress((idx + 1) / len(archivos_subidos), text=f"✓ {archivo.name}")

    progress.progress(1.0, text="¡Completado!")

    if todos_los_comentarios:
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#1B3A6B,#2C5F8A);
                    border-radius:10px; padding:20px; text-align:center; margin:20px 0">
            <h2 style="color:white; margin:0">
                {len(todos_los_comentarios)} comentarios procesados
            </h2>
            <p style="color:#A8C8E8; margin:5px 0 0 0">
                de {len(archivos_subidos)} documento(s)
            </p>
        </div>
        """, unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        if "Excel" in output_elegido:
            with col1:
                excel_bytes = generar_excel_bytes(todos_los_comentarios)
                st.download_button(
                    label="📥 Descargar Excel",
                    data=excel_bytes,
                    file_name="comentarios.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        if "PDF" in output_elegido:
            with col2:
                pdf_bytes = generar_pdf_bytes(todos_los_comentarios)
                st.download_button(
                    label="📥 Descargar PDF",
                    data=pdf_bytes,
                    file_name="comentarios.pdf",
                    mime="application/pdf"
                )
    else:
        st.info("ℹ️ No se encontraron comentarios en los archivos procesados.")

# Footer
st.markdown("""
<div class="caf-footer">
    <span>CAF</span> · Banco de Desarrollo de América Latina y el Caribe<br>
    Dirección de Riesgo Soberano
</div>
""", unsafe_allow_html=True)
